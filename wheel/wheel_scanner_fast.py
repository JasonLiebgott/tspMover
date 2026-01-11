"""
Fast Wheel Strategy Scanner
Optimized for speed and practical use
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
import warnings
import time
warnings.filterwarnings('ignore')


def calculate_rsi_fast(prices, period=14):
    """Fast RSI calculation"""
    delta = prices.diff()
    gain = delta.where(delta > 0, 0).rolling(window=period).mean()
    loss = -delta.where(delta < 0, 0).rolling(window=period).mean()
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    return rsi.iloc[-1] if len(rsi) > 0 and not pd.isna(rsi.iloc[-1]) else 50


def scan_ticker_fast(ticker, min_dte=15, max_dte=60, max_collateral=25000):
    """Fast scan of a single ticker"""
    
    try:
        stock = yf.Ticker(ticker)
        
        # Get current price
        current_price = stock.info.get('currentPrice', stock.info.get('regularMarketPrice', 0))
        if current_price <= 0:
            return []
        
        # Calculate RSI quickly
        try:
            hist = stock.history(period='1mo')
            rsi = calculate_rsi_fast(hist['Close'], 14) if len(hist) > 20 else 50
        except:
            rsi = 50
        
        # Get expirations
        expirations = list(stock.options)
        if len(expirations) == 0:
            return []
        
        candidates = []
        today = datetime.now()
        
        # Scan first 6 expirations only
        for exp_str in expirations[:6]:
            try:
                exp_date = datetime.strptime(exp_str, '%Y-%m-%d')
                days = (exp_date - today).days
                
                if days < min_dte or days > max_dte:
                    continue
                
                # Get puts only
                chain = stock.option_chain(exp_str)
                puts = chain.puts
                
                # Filter OTM puts quickly
                otm_puts = puts[
                    (puts['strike'] < current_price) &  # OTM
                    (puts['bid'] > 0) &  # Has bid
                    (puts['ask'] > 0) &  # Has ask
                    (puts['strike'] * 100 <= max_collateral)  # Affordable
                ]
                
                for idx, row in otm_puts.iterrows():
                    strike = row['strike']
                    bid = row['bid']
                    ask = row['ask']
                    volume = row.get('volume', 0)
                    oi = row.get('openInterest', 0)
                    delta = row.get('delta', -0.25)
                    iv = row.get('impliedVolatility', 0)
                    
                    # Calculate metrics
                    mid = (bid + ask) / 2
                    premium = mid * 100
                    collateral = strike * 100
                    roc = (premium / collateral) * 100
                    annual = (roc / days) * 365
                    spread_pct = ((ask - bid) / mid) * 100 if mid > 0 else 100
                    cushion = ((current_price - strike) / current_price) * 100
                    pop = (1 - abs(delta)) * 100
                    ppd = premium / days
                    
                    # Basic filters
                    if roc < 0.2:  # Less than 0.2% ROC
                        continue
                    if annual < 5:  # Less than 5% annualized
                        continue
                    if spread_pct > 30:  # Spread too wide
                        continue
                    if pop < 50:  # Less than 50% PoP
                        continue
                    
                    candidates.append({
                        'ticker': ticker,
                        'expiry': exp_str,
                        'dte': days,
                        'strike': strike,
                        'delta': delta,
                        'premium': premium,
                        'bid': bid,
                        'ask': ask,
                        'iv': iv * 100,
                        'roc': roc,
                        'annual_yield': annual,
                        'pop': pop,
                        'spread_pct': spread_pct,
                        'cushion': cushion,
                        'rsi': rsi,
                        'collateral': collateral,
                        'ppd': ppd,
                        'volume': volume,
                        'oi': oi,
                        'price': current_price,
                    })
            
            except:
                continue
        
        return candidates
    
    except Exception as e:
        return []


def scan_universe(tickers, max_collateral=25000):
    """Scan list of tickers"""
    
    print(f"\n{'='*120}")
    print(f"FAST WHEEL STRATEGY SCANNER")
    print(f"Scanning {len(tickers)} tickers")
    print(f"Max Collateral: ${max_collateral:,}")
    print(f"{'='*120}\n")
    
    all_candidates = []
    
    for i, ticker in enumerate(tickers, 1):
        # Add small delay to avoid rate limiting (skip delay for first ticker)
        if i > 1:
            time.sleep(0.5)  # 500ms delay between tickers
        
        print(f"[{i}/{len(tickers)}] {ticker}...", end=' ', flush=True)
        candidates = scan_ticker_fast(ticker, max_collateral=max_collateral)
        if len(candidates) > 0:
            print(f"✓ {len(candidates)} found")
            all_candidates.extend(candidates)
        else:
            print("—")
    
    if len(all_candidates) == 0:
        print("\nNo candidates found.")
        return pd.DataFrame()
    
    df = pd.DataFrame(all_candidates)
    df = df.sort_values(['annual_yield', 'pop'], ascending=[False, False])
    
    return df


def display_results(df, top_n=30):
    """Display formatted results with risk filtering"""
    
    if len(df) == 0:
        return
    
    # Calculate risk score for filtering
    df_display = df.copy()
    df_display['risk_score'] = 0
    
    # Add risk penalties
    df_display.loc[df_display['oi'] < 50, 'risk_score'] += 3
    df_display.loc[df_display['volume'] < 10, 'risk_score'] += 2
    df_display.loc[df_display['spread_pct'] > 15, 'risk_score'] += 2
    df_display.loc[df_display['roc'] < 0.5, 'risk_score'] += 2
    df_display.loc[df_display['cushion'] < 2, 'risk_score'] += 2
    df_display.loc[df_display['dte'] < 21, 'risk_score'] += 1
    df_display.loc[df_display['annual_yield'] > 50, 'risk_score'] += 1
    df_display.loc[df_display['delta'].abs() > 0.30, 'risk_score'] += 1
    
    # Assign risk levels
    df_display['risk_level'] = 'RED'
    df_display.loc[df_display['risk_score'] <= 3, 'risk_level'] = 'YELLOW'
    df_display.loc[df_display['risk_score'] == 0, 'risk_level'] = 'GREEN'
    
    # Count by risk level
    green_count = len(df_display[df_display['risk_level'] == 'GREEN'])
    yellow_count = len(df_display[df_display['risk_level'] == 'YELLOW'])
    red_count = len(df_display[df_display['risk_level'] == 'RED'])
    
    print(f"\n{'='*120}")
    print(f"CASH-SECURED PUT CANDIDATES - RISK FILTERED")
    print(f"{'='*120}")
    print(f"GREEN (Best): {green_count} | YELLOW (Acceptable): {yellow_count} | RED (Avoid): {red_count}")
    print(f"{'='*120}\n")
    
    # Show GREEN candidates first
    if green_count > 0:
        print(f"*** GREEN - BEST CANDIDATES (Good Liquidity, Conservative, Easy Exit) ***\n")
        green_df = df_display[df_display['risk_level'] == 'GREEN'].head(20)
    
    # Show YELLOW candidates
    if yellow_count > 0:
        print(f"\n*** YELLOW - ACCEPTABLE WITH CAUTION (Monitor closely, close early) ***\n")
        yellow_df = df_display[df_display['risk_level'] == 'YELLOW'].head(15)
        
        output = pd.DataFrame({
            'Ticker': yellow_df['ticker'],
            'Expiry': yellow_df['expiry'],
            'DTE': yellow_df['dte'].astype(int),
            'Strike': yellow_df['strike'].map('${:.2f}'.format),
            'Δ': yellow_df['delta'].map('{:.3f}'.format),
            'Prem': yellow_df['premium'].map('${:.0f}'.format),
            'ROC%': yellow_df['roc'].map('{:.2f}'.format),
            'Annual%': yellow_df['annual_yield'].map('{:.1f}'.format),
            'PoP%': yellow_df['pop'].map('{:.0f}'.format),
            'Cushion%': yellow_df['cushion'].map('{:.1f}'.format),
            'Spread%': yellow_df['spread_pct'].map('{:.1f}'.format),
            'Vol': yellow_df['volume'].astype(int),
            'OI': yellow_df['oi'].astype(int),
        })
        print(output.to_string(index=False))
    
    # Don't show RED candidates - we're avoiding them
    if red_count > 0:
        print(f"\n*** {red_count} RED (High Risk) candidates hidden - see Excel for details ***")
    
    # Summary stats
    print(f"\n{'='*120}")
    print("SUMMARY STATISTICS")
    print(f"{'='*120}\n")
    
    print(f"Total Candidates: {len(df)}")
    print(f"Risk Breakdown: {green_count} GREEN | {yellow_count} YELLOW | {red_count} RED")
    print(f"Unique Tickers: {df['ticker'].nunique()}")
    print(f"\nAverage Metrics (All Candidates):")
    print(f"  ROC: {df['roc'].mean():.2f}% (range: {df['roc'].min():.2f}% - {df['roc'].max():.2f}%)")
    print(f"  Annual Yield: {df['annual_yield'].mean():.1f}% (range: {df['annual_yield'].min():.1f}% - {df['annual_yield'].max():.1f}%)")
    print(f"  PoP: {df['pop'].mean():.1f}%")
    print(f"  Days to Expiry: {df['dte'].mean():.0f} days")
    print(f"  Cushion: {df['cushion'].mean():.1f}%")
    
    if green_count > 0:
        green_df_stats = df_display[df_display['risk_level'] == 'GREEN']
        print(f"\nGREEN Candidates Only:")
        print(f"  Average ROC: {green_df_stats['roc'].mean():.2f}%")
        print(f"  Average Annual: {green_df_stats['annual_yield'].mean():.1f}%")
        print(f"  Average Cushion: {green_df_stats['cushion'].mean():.1f}%")
        print(f"  Average OI: {green_df_stats['oi'].mean():.0f}")
        
        print(f"\nTop 5 GREEN Candidates:")
        top5 = green_df_stats.nlargest(5, 'annual_yield')[['ticker', 'strike', 'dte', 'annual_yield', 'roc', 'cushion', 'oi']]
        for idx, row in top5.iterrows():
            print(f"  {row['ticker']} ${row['strike']:.2f} ({row['dte']}d): {row['annual_yield']:.1f}% annual, {row['roc']:.2f}% ROC, {row['cushion']:.1f}% cushion, OI:{row['oi']:.0f}")
    
    print(f"\nRECOMMENDATION:")
    if green_count == 0:
        print("  NO TRADES - Wait for better opportunities with good liquidity and conservative positioning")
    elif green_count < 5:
        print(f"  LIMITED OPPORTUNITIES - Only {green_count} low-risk candidates available. Be selective.")
    else:
        print(f"  GOOD ENVIRONMENT - {green_count} quality candidates. Focus on those with highest OI for easy exits.")


def save_results(df, filename=None):
    """Save to Excel with color coding and risk highlighting"""
    if len(df) == 0:
        return
    
    if filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"wheel_scan_{timestamp}.xlsx"
    
    # Create Excel writer with formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Wheel Candidates', index=False)
        
        # Get workbook and worksheet
        from openpyxl.styles import PatternFill, Font
        from openpyxl.formatting.rule import CellIsRule
        
        workbook = writer.book
        worksheet = writer.sheets['Wheel Candidates']
        
        # Define color fills
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red
        bold_font = Font(bold=True)
        
        # Color code entire rows based on risk criteria
        for idx, row_data in df.iterrows():
            row_num = idx + 2  # Excel rows start at 1, header is row 1
            
            # Risk scoring (lower is better)
            risk_score = 0
            
            # Liquidity check (critical for closing early)
            if row_data['oi'] < 50:
                risk_score += 3  # Thin liquidity - hard to exit
            if row_data['volume'] < 10:
                risk_score += 2
            if row_data['spread_pct'] > 15:
                risk_score += 2  # Wide spread - hard to get good fill
            
            # ROC too low - not worth the risk
            if row_data['roc'] < 0.5:
                risk_score += 2
            
            # Cushion too thin - high assignment risk
            if row_data['cushion'] < 2:
                risk_score += 2
            
            # Days too short - less time to close early
            if row_data['dte'] < 21:
                risk_score += 1
            
            # Annual yield too high might indicate hidden risk
            if row_data['annual_yield'] > 50:
                risk_score += 1
            
            # Delta too high - more likely to get assigned
            if abs(row_data['delta']) > 0.30:
                risk_score += 1
            
            # Apply color to entire row based on risk
            if risk_score == 0:
                # GREEN: Best candidates - good liquidity, conservative, easy to exit
                fill = green_fill
            elif risk_score <= 3:
                # YELLOW: Acceptable but monitor closely
                fill = yellow_fill
            else:
                # RED: High risk - avoid or be very careful
                fill = red_fill
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col)
                cell.fill = fill
        
        # Bold the header row
        for cell in worksheet[1]:
            cell.font = bold_font
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add legend sheet
        legend_df = pd.DataFrame({
            'Risk Level': ['GREEN', 'YELLOW', 'RED'],
            'Criteria': [
                'BEST: Good liquidity (OI>50, Vol>10), Conservative (Cushion>2%, ROC>0.5%), Easy to exit (Spread<15%)',
                'ACCEPTABLE: Some concerns but tradeable, monitor closely for early exit opportunities',
                'AVOID: Thin liquidity, wide spreads, or too aggressive - difficult to close early, high assignment risk'
            ],
            'Action': [
                'Trade with confidence, set profit target at 50% of premium to close early',
                'Trade cautiously, close at 40% profit or if uncomfortable',
                'Avoid unless you want assignment. If trading, use tight stops and close quickly'
            ]
        })
        legend_df.to_excel(writer, sheet_name='Risk Legend', index=False)
        
        # Format legend
        legend_ws = writer.sheets['Risk Legend']
        legend_ws.cell(2, 1).fill = green_fill
        legend_ws.cell(3, 1).fill = yellow_fill
        legend_ws.cell(4, 1).fill = red_fill
        
        for cell in legend_ws[1]:
            cell.font = bold_font
        
        for column in legend_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            legend_ws.column_dimensions[column_letter].width = adjusted_width
    
    print(f"\nResults saved to: {filename}")
    print("\nColor Coding:")
    print("  GREEN  = Best candidates - good liquidity, easy to close early for profit")
    print("  YELLOW = Acceptable with caution - monitor and close at 40-50% profit")
    print("  RED    = High risk - thin liquidity or aggressive, avoid unless you want assignment")


def main():
    """Main execution"""
    
    # Define universe - consolidated from all scanners
    tickers = [
        # Large cap tech
        'AAPL', 'MSFT', 'GOOGL', 'META', 'NVDA', 'AMD', 'INTC', 'CSCO', 'PLTR', 'ANET', 'HOOD',
        # Finance
        'JPM', 'BAC', 'WFC', 'C', 'GS', 'MS', 'SCHW',
        # Consumer
        'KO', 'PEP', 'WMT', 'TGT', 'HD', 'LOW', 'MCD', 'SBUX', 'NKE', 'DIS',
        # Healthcare
        'JNJ', 'PFE', 'UNH', 'CVS', 'ABBV', 'MRK', 'LLY',
        # Energy
        'XOM', 'CVX', 'COP', 'SLB', 'HAL',
        # Industrial
        'BA', 'CAT', 'GE', 'MMM', 'HON',
        # Telecom
        'T', 'VZ', 'TMUS',
        # Consumer staples
        'PG', 'KMB', 'CL',
        # ETFs
        'SPY', 'QQQ', 'IWM', 'DIA', 'XLF', 'XLE', 'XLK'
    ]
    
    # Scan
    results = scan_universe(tickers, max_collateral=25000)
    
    # Display
    display_results(results, top_n=30)
    
    # Save
    if len(results) > 0:
        save_results(results)


if __name__ == "__main__":
    main()
