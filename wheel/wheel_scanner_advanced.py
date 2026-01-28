"""
Wheel Strategy Scanner - Advanced Options Screener
Scans option universe for cash-secured puts and covered calls
with comprehensive metrics and filtering
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
import warnings
import time
import os
import smtplib
from email.message import EmailMessage
from zoneinfo import ZoneInfo
import argparse
from scipy.stats import norm
from scipy.stats import percentileofscore
warnings.filterwarnings('ignore')


class DataProvider:
    """Base class for data providers - pluggable architecture"""
    
    def get_current_price(self, ticker: str) -> float:
        raise NotImplementedError
    
    def get_option_chain(self, ticker: str, expiration: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Returns (calls_df, puts_df)"""
        raise NotImplementedError
    
    def get_ohlcv_data(self, ticker: str, period: str = "3mo") -> pd.DataFrame:
        raise NotImplementedError
    
    def get_expirations(self, ticker: str) -> List[str]:
        raise NotImplementedError


class YFinanceProvider(DataProvider):
    """Yahoo Finance data provider"""
    
    def get_current_price(self, ticker: str) -> float:
        stock = yf.Ticker(ticker)
        info = stock.info
        return info.get('currentPrice', info.get('regularMarketPrice', 0))
    
    def get_option_chain(self, ticker: str, expiration: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        stock = yf.Ticker(ticker)
        chain = stock.option_chain(expiration)
        return chain.calls, chain.puts
    
    def get_ohlcv_data(self, ticker: str, period: str = "3mo") -> pd.DataFrame:
        stock = yf.Ticker(ticker)
        return stock.history(period=period)
    
    def get_expirations(self, ticker: str) -> List[str]:
        stock = yf.Ticker(ticker)
        return list(stock.options)
    
    def get_earnings_date(self, ticker: str) -> Optional[datetime]:
        """Get next earnings date for ticker, returns None if not available"""
        try:
            stock = yf.Ticker(ticker)
            # Use calendar instead of earnings_dates (faster and more reliable)
            calendar = stock.calendar
            if calendar is not None and 'Earnings Date' in calendar:
                earnings_date = calendar['Earnings Date']
                # Handle list of earnings dates (take the first one)
                if isinstance(earnings_date, list) and len(earnings_date) > 0:
                    earnings_date = earnings_date[0]
                # Convert to datetime
                if isinstance(earnings_date, pd.Timestamp):
                    return earnings_date.tz_localize(None) if earnings_date.tz else earnings_date.to_pydatetime()
                elif isinstance(earnings_date, datetime):
                    return earnings_date
                else:
                    # Convert date object to datetime
                    from datetime import date
                    if isinstance(earnings_date, date):
                        return datetime.combine(earnings_date, datetime.min.time())
            return None
        except Exception as e:
            # Don't let earnings fetch failure stop the entire scan
            return None
    
    def get_dividend_info(self, ticker: str) -> Tuple[Optional[datetime], float]:
        """Get next ex-dividend date and amount, returns (None, 0.0) if not available"""
        try:
            stock = yf.Ticker(ticker)
            # Get dividend info from calendar
            calendar = stock.calendar
            if calendar is not None:
                ex_div_date = calendar.get('Ex-Dividend Date')
                dividend_date = calendar.get('Dividend Date')
                
                # Use ex-dividend date (when stock price drops)
                if ex_div_date is not None:
                    if isinstance(ex_div_date, list) and len(ex_div_date) > 0:
                        ex_div_date = ex_div_date[0]
                    if hasattr(ex_div_date, 'date'):
                        from datetime import date
                        if isinstance(ex_div_date, date):
                            ex_div_date = datetime.combine(ex_div_date, datetime.min.time())
                    
                    # Get dividend amount from dividends history
                    dividends = stock.dividends
                    if dividends is not None and len(dividends) > 0:
                        # Use most recent dividend amount
                        div_amount = dividends.iloc[-1]
                        return ex_div_date, float(div_amount)
            
            return None, 0.0
        except Exception:
            return None, 0.0
    
    def get_stock_risk_metrics(self, ticker: str) -> Dict[str, float]:
        """Get beta, market cap, and historical volatility for stock risk assessment"""
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            
            # Get beta (vs SPY/market)
            beta = info.get('beta', None)
            if beta is None or pd.isna(beta):
                beta = 1.0  # Assume market beta if unknown
            
            # Get market cap
            market_cap = info.get('marketCap', 0)
            
            # Calculate historical volatility (30-day realized)
            hist = stock.history(period='1mo')
            if len(hist) > 5:
                returns = hist['Close'].pct_change().dropna()
                hist_vol = returns.std() * np.sqrt(252)  # Annualized
            else:
                hist_vol = 0.30  # Default assumption
            
            # Calculate Average True Range (% of price)
            if len(hist) > 14:
                high_low = hist['High'] - hist['Low']
                high_close = abs(hist['High'] - hist['Close'].shift())
                low_close = abs(hist['Low'] - hist['Close'].shift())
                tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
                atr = tr.rolling(14).mean().iloc[-1]
                atr_pct = (atr / hist['Close'].iloc[-1]) * 100  # ATR as % of price
            else:
                atr_pct = 2.0  # Default assumption
            
            return {
                'beta': float(beta),
                'market_cap': float(market_cap),
                'hist_vol': float(hist_vol),
                'atr_pct': float(atr_pct)
            }
        except Exception:
            # Return conservative defaults if fetch fails
            return {
                'beta': 1.0,
                'market_cap': 0,
                'hist_vol': 0.30,
                'atr_pct': 2.0
            }


class BlackScholesCalculator:
    """Black-Scholes option pricing and Greeks calculations"""
    
    @staticmethod
    def calculate_delta(option_type: str, spot_price: float, strike: float, 
                       time_to_expiry_days: float, iv: float, risk_free_rate: float = 0.05) -> float:
        """
        Calculate delta using Black-Scholes formula
        
        Args:
            option_type: 'call' or 'put'
            spot_price: Current stock price
            strike: Option strike price
            time_to_expiry_days: Days until expiration
            iv: Implied volatility (as decimal, e.g., 0.25 for 25%)
            risk_free_rate: Annual risk-free rate (default 5%)
        
        Returns:
            Delta value (call: 0 to 1, put: -1 to 0)
        """
        if iv <= 0 or time_to_expiry_days <= 0:
            # Fallback to moneyness-based estimation if IV is missing
            moneyness = strike / spot_price
            if option_type == 'put':
                if moneyness > 0.95:
                    return -0.40
                elif moneyness > 0.90:
                    return -0.25
                else:
                    return -0.15
            else:  # call
                if moneyness < 1.05:
                    return 0.40
                elif moneyness < 1.10:
                    return 0.25
                else:
                    return 0.15
        
        try:
            # Convert days to years
            T = time_to_expiry_days / 365.0
            
            # Calculate d1
            d1 = (np.log(spot_price / strike) + (risk_free_rate + 0.5 * iv**2) * T) / (iv * np.sqrt(T))
            
            # Calculate delta
            if option_type == 'call':
                delta = norm.cdf(d1)
            else:  # put
                delta = norm.cdf(d1) - 1
            
            return delta
        
        except Exception:
            # Fallback to moneyness estimation if calculation fails
            moneyness = strike / spot_price
            if option_type == 'put':
                if moneyness > 0.95:
                    return -0.40
                elif moneyness > 0.90:
                    return -0.25
                else:
                    return -0.15
            else:  # call
                if moneyness < 1.05:
                    return 0.40
                elif moneyness < 1.10:
                    return 0.25
                else:
                    return 0.15
    
    @staticmethod
    def calculate_probability_otm(option_type: str, spot_price: float, strike: float,
                                  time_to_expiry_days: float, iv: float, risk_free_rate: float = 0.05) -> float:
        """
        Calculate probability that option expires OTM (out-of-the-money) using Black-Scholes d2.
        This is the PROPER probability calculation for short options, not the delta approximation.
        
        For short puts: Probability stock stays ABOVE strike (put expires worthless)
        For short calls: Probability stock stays BELOW strike (call expires worthless)
        
        Args:
            option_type: 'call' or 'put'
            spot_price: Current stock price
            strike: Option strike price
            time_to_expiry_days: Days until expiration
            iv: Implied volatility (as decimal, e.g., 0.25 for 25%)
            risk_free_rate: Annual risk-free rate (default 5%)
        
        Returns:
            Probability option expires OTM (0.0 to 1.0)
        """
        if iv <= 0 or time_to_expiry_days <= 0:
            # Fallback to moneyness-based estimation if IV is missing
            moneyness = strike / spot_price
            if option_type == 'put':
                if moneyness > 0.95:  # Close to ATM
                    return 0.60
                elif moneyness > 0.90:  # Moderately OTM
                    return 0.75
                else:  # Deep OTM
                    return 0.85
            else:  # call
                if moneyness < 1.05:  # Close to ATM
                    return 0.60
                elif moneyness < 1.10:  # Moderately OTM
                    return 0.75
                else:  # Deep OTM
                    return 0.85
        
        try:
            # Convert days to years
            T = time_to_expiry_days / 365.0
            
            # Calculate d1 and d2
            d1 = (np.log(spot_price / strike) + (risk_free_rate + 0.5 * iv**2) * T) / (iv * np.sqrt(T))
            d2 = d1 - iv * np.sqrt(T)
            
            # Calculate probability based on option type
            # d2 represents the probability of finishing ITM under risk-neutral measure
            # We want probability of finishing OTM (what we sell expires worthless)
            if option_type == 'put':
                # For short put: probability stock stays ABOVE strike
                prob_otm = norm.cdf(d2)
            else:  # call
                # For short call: probability stock stays BELOW strike
                prob_otm = norm.cdf(-d2)
            
            return max(0.0, min(1.0, prob_otm))  # Clamp to [0, 1]
        
        except Exception:
            # Fallback to moneyness estimation if calculation fails
            moneyness = strike / spot_price
            if option_type == 'put':
                if moneyness > 0.95:
                    return 0.60
                elif moneyness > 0.90:
                    return 0.75
                else:
                    return 0.85
            else:  # call
                if moneyness < 1.05:
                    return 0.60
                elif moneyness < 1.10:
                    return 0.75
                else:
                    return 0.85


class TechnicalIndicators:
    """Calculate technical indicators manually if pandas_ta not available"""
    
    @staticmethod
    def calculate_rsi(prices: pd.Series, period: int = 14) -> float:
        """Calculate RSI indicator"""
        delta = prices.diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
        
        # Prevent division by zero
        rs = gain / loss.replace(0, 1e-10)
        rsi = 100 - (100 / (1 + rs))
        return rsi.iloc[-1]
    
    @staticmethod
    def calculate_adx(df: pd.DataFrame, period: int = 14) -> float:
        """Calculate ADX indicator"""
        high = df['High']
        low = df['Low']
        close = df['Close']
        
        # True Range
        tr1 = high - low
        tr2 = abs(high - close.shift())
        tr3 = abs(low - close.shift())
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
        
        # Directional Movement
        up_move = high - high.shift()
        down_move = low.shift() - low
        
        plus_dm = np.where((up_move > down_move) & (up_move > 0), up_move, 0)
        minus_dm = np.where((down_move > up_move) & (down_move > 0), down_move, 0)
        
        # Smoothed indicators
        atr = tr.rolling(window=period).mean()
        plus_di = 100 * pd.Series(plus_dm).rolling(window=period).mean() / atr
        minus_di = 100 * pd.Series(minus_dm).rolling(window=period).mean() / atr
        
        # ADX - prevent division by zero
        denominator = (plus_di + minus_di).replace(0, 1e-10)
        dx = 100 * abs(plus_di - minus_di) / denominator
        adx = dx.rolling(window=period).mean()
        
        return adx.iloc[-1] if not pd.isna(adx.iloc[-1]) else 0


class WheelConfig:
    """Configuration for wheel strategy screening"""
    
    def __init__(self):
        # Days to expiration range
        self.min_dte = 15
        self.max_dte = 60
        
        # Delta ranges - relaxed to match fast scanner
        self.min_put_delta = 0.05  # More lenient
        self.max_put_delta = 0.50  # More lenient
        self.min_call_delta = 0.05
        self.max_call_delta = 0.50
        
        # Return thresholds
        self.min_roc = 0.002  # 0.2% minimum return on capital
        self.min_annualized = 0.05  # 5% minimum annualized
        
        # Liquidity filters
        self.max_spread_pct = 0.10  # 10% max bid-ask spread (tightened from 30%)
        self.min_volume = 10  # Minimum 10 contracts traded (OR condition with OI)
        self.min_open_interest = 200  # Minimum 200 OI (OR condition with volume)
        
        # Probability and cushion (80-85% PoP is ideal)
        self.min_pop = 0.70  # 70% minimum probability of profit (allows 70-85% range)
        self.min_cushion = 0.00  # Require OTM puts (0% minimum cushion for CSP)
        self.min_call_cushion = -0.02  # Allow slightly ITM for CC (2% ITM acceptable)
        
        # Technical filters
        self.max_rsi_for_csp = 85  # More lenient - allow overbought
        self.min_rsi_for_csp = 15  # Allow oversold
        self.max_adx = 70  # Allow strong trends
        
        # Capital constraints
        self.max_collateral = 25000  # Maximum collateral per position
        
        # BORING WHEEL FILTERS (for true conservative wheel strategy)
        self.max_beta = 1.2  # Maximum beta (volatility vs market) - 1.0 = market, <1 = defensive
        self.max_hist_vol = 0.40  # Maximum 40% historical volatility (annualized)
        self.max_atr_pct = 3.5  # Maximum 3.5% daily price swing (ATR as % of price)
        self.min_market_cap = 10_000_000_000  # $10B minimum (large cap only)
        self.max_annualized_for_boring = 0.25  # 25% max annualized (anything higher = hidden risk)
        self.boring_wheel_mode = False  # Set to True to STRICTLY reject non-boring stocks
        # NOTE: When False, runs "mixed mode" - boring-friendly scoring but allows non-boring stocks with penalty
        
        # HIGH-VOL / GAP-PRONE STOCKS: Apply stricter rules
        self.high_vol_tickers = {'ABNB', 'PINS', 'WDAY', 'ZS', 'PANW', 'FTNT'}
        self.high_vol_min_cushion = 0.15  # 15% minimum cushion (vs -2% normal)
        self.high_vol_min_pop = 0.85  # 85% minimum PoP (vs 70% normal)
        self.high_vol_min_dte = 21  # 21 days minimum (vs 15 normal)
        self.high_vol_max_dte = 45  # 45 days maximum (vs 60 normal)
        self.high_vol_exclude_earnings = True  # Exclude if earnings before expiry
        
        # EARNINGS RISK MANAGEMENT (applies to ALL tickers, not just high-vol)
        self.exclude_earnings_all = False  # Set True to exclude ALL options with earnings before expiry
        self.earnings_penalty_in_scoring = True  # Add risk penalty for earnings (if not excluding)

        # Scheduling (market open window) and local timezone
        # NOTE: If you want true MST year-round (no DST), use "America/Phoenix"
        self.user_timezone = "America/Denver"
        self.market_open_window_minutes = 15

        # GO/NO GO thresholds for top boring candidate
        self.go_min_composite_score = 70.0
        self.go_min_pop = 0.80
        self.go_min_cushion = 0.03


class WheelScanner:
    """Main scanner for wheel strategy candidates"""
    
    def __init__(self, data_provider: DataProvider, config: WheelConfig, debug: bool = False):
        self.provider = data_provider
        self.config = config
        self.current_date = datetime.now()
        self.debug = debug
        self.earnings_cache = {}  # Cache earnings dates to avoid repeated API calls
        self.dividend_cache = {}  # Cache dividend info to avoid repeated API calls
        self.risk_metrics_cache = {}  # Cache beta, volatility, ATR to avoid repeated API calls
        self.filter_stats = {
            'total_options': 0,
            'dte_reject': 0,
            'otm_reject': 0,
            'delta_reject': 0,
            'cushion_reject': 0,
            'rsi_reject': 0,
            'collateral_reject': 0,
            'roc_reject': 0,
            'annualized_reject': 0,
            'pop_reject': 0,
            'spread_reject': 0,
            'liquidity_reject': 0,
            'passed': 0
        }
    
    def calculate_technicals(self, ticker: str) -> Dict[str, float]:
        """Calculate RSI and ADX for a ticker using manual calculations"""
        try:
            df = self.provider.get_ohlcv_data(ticker)
            
            if len(df) < 30:
                return {'rsi': 50, 'adx': 20}  # Defaults
            
            # Manual calculation (always use these)
            rsi = TechnicalIndicators.calculate_rsi(df['Close'], 14)
            adx = TechnicalIndicators.calculate_adx(df, 14)
            
            return {'rsi': rsi, 'adx': adx}
        
        except Exception as e:
            return {'rsi': 50, 'adx': 20}  # Return defaults on error
    
    def analyze_option(self, ticker: str, row: pd.Series, option_type: str, 
                      underlying_price: float, expiry_date: datetime,
                      technicals: Dict[str, float]) -> Optional[Dict]:
        """Analyze a single option and return metrics if it passes filters"""
        
        try:
            # Basic data
            strike = row['strike']
            bid = row.get('bid', 0)
            ask = row.get('ask', 0)
            last = row.get('lastPrice', 0)
            volume = row.get('volume', 0)
            open_interest = row.get('openInterest', 0)
            delta = row.get('delta', 0)
            iv = row.get('impliedVolatility', 0)
            
            # Skip if no bid/ask
            if bid <= 0 or ask <= 0:
                return None
            
            # Days to expiry (use trading days, not calendar days)
            days_to_expiry = np.busday_count(self.current_date.date(), expiry_date.date())
            # Ensure at least 1 day to avoid division by zero
            if days_to_expiry < 1:
                days_to_expiry = 1
            
            # Check if this is a high-vol/gap-prone stock
            is_high_vol_stock = ticker in self.config.high_vol_tickers
            
            # Filter by DTE (stricter for high-vol stocks)
            if is_high_vol_stock:
                if days_to_expiry < self.config.high_vol_min_dte or days_to_expiry > self.config.high_vol_max_dte:
                    return None
            else:
                if days_to_expiry < self.config.min_dte or days_to_expiry > self.config.max_dte:
                    return None
            
            # Mid price and premium
            mid_price = (bid + ask) / 2
            premium_per_contract = mid_price * 100
            
            # Determine wheel leg type and calculate metrics
            if option_type == 'put':
                # Cash-secured put
                leg_type = 'CSP'
                collateral = strike * 100
                
                # Check for dividend during option period and adjust price
                if ticker not in self.dividend_cache:
                    self.dividend_cache[ticker] = self.provider.get_dividend_info(ticker)
                
                ex_div_date, div_amount = self.dividend_cache[ticker]
                adjusted_price = underlying_price
                dividend_risk = False
                
                if ex_div_date and div_amount > 0:
                    # If ex-dividend is during option period, adjust price
                    if self.current_date < ex_div_date < expiry_date:
                        adjusted_price = underlying_price - div_amount
                        dividend_risk = True
                
                # Calculate cushion with dividend-adjusted price
                cushion = (adjusted_price - strike) / adjusted_price
                
                # Must be OTM - check against dividend-adjusted price
                if strike >= adjusted_price:
                    return None
                
                # If delta is missing or 0, calculate it using Black-Scholes
                if delta == 0 or pd.isna(delta):
                    delta = BlackScholesCalculator.calculate_delta(
                        option_type='put',
                        spot_price=underlying_price,
                        strike=strike,
                        time_to_expiry_days=days_to_expiry,
                        iv=iv if iv > 0 else 0.30  # Use 30% default IV if missing
                    )
                
                # Filter by delta
                if abs(delta) < self.config.min_put_delta or abs(delta) > self.config.max_put_delta:
                    return None
                
                # Filter by cushion (stricter for high-vol stocks)
                min_cushion_threshold = self.config.high_vol_min_cushion if is_high_vol_stock else self.config.min_cushion
                if cushion < min_cushion_threshold:
                    return None
                
                # PoP for short put - Use proper Black-Scholes d2 probability
                # This is probability stock stays ABOVE strike (put expires worthless)
                pop = BlackScholesCalculator.calculate_probability_otm(
                    option_type='put',
                    spot_price=underlying_price,
                    strike=strike,
                    time_to_expiry_days=days_to_expiry,
                    iv=iv if iv > 0 else 0.30  # Use 30% default IV if missing
                )
                
                # Technical filter: avoid overbought for CSP
                if technicals['rsi'] > self.config.max_rsi_for_csp:
                    return None
                
            else:  # call
                # Covered call
                leg_type = 'CC'
                collateral = underlying_price * 100
                cushion = (strike - underlying_price) / underlying_price
                
                # Prefer OTM but allow slightly ITM
                if cushion < -0.02:  # More than 2% ITM, skip
                    return None
                
                # If delta is missing or 0, calculate it using Black-Scholes
                if delta == 0 or pd.isna(delta):
                    delta = BlackScholesCalculator.calculate_delta(
                        option_type='call',
                        spot_price=underlying_price,
                        strike=strike,
                        time_to_expiry_days=days_to_expiry,
                        iv=iv if iv > 0 else 0.30  # Use 30% default IV if missing
                    )
                
                # Filter by delta
                if abs(delta) < self.config.min_call_delta or abs(delta) > self.config.max_call_delta:
                    return None
                
                # PoP for short call - Use proper Black-Scholes d2 probability
                # This is probability stock stays BELOW strike (call expires worthless)
                pop = BlackScholesCalculator.calculate_probability_otm(
                    option_type='call',
                    spot_price=underlying_price,
                    strike=strike,
                    time_to_expiry_days=days_to_expiry,
                    iv=iv if iv > 0 else 0.30  # Use 30% default IV if missing
                )
                
                # Technical filter: avoid strong uptrends for CC
                if technicals['adx'] > self.config.max_adx and cushion > 0:
                    return None
            
            # Filter by collateral
            if collateral > self.config.max_collateral:
                return None
            
            # Return on capital
            roc = premium_per_contract / collateral
            
            # Filter by ROC
            if roc < self.config.min_roc:
                return None
            
            # Annualized yield (use 252 trading days to match trading-day DTE)
            annualized_yield = roc * (252 / days_to_expiry)
            
            # Filter by annualized yield
            if annualized_yield < self.config.min_annualized:
                return None
            
            # Filter by PoP (stricter for high-vol stocks)
            min_pop_threshold = self.config.high_vol_min_pop if is_high_vol_stock else self.config.min_pop
            if pop < min_pop_threshold:
                return None
            
            # Bid-ask spread
            spread_abs = ask - bid
            spread_pct = spread_abs / mid_price if mid_price > 0 else 1
            
            # Filter by spread
            if spread_pct > self.config.max_spread_pct:
                return None
            
            # Filter by liquidity (option must have EITHER sufficient volume OR open interest)
            # Only reject if BOTH volume AND open interest are insufficient
            has_volume = volume >= self.config.min_volume
            has_open_interest = open_interest >= self.config.min_open_interest
            if not (has_volume or has_open_interest):
                return None
            
            # Profit per day
            ppd = premium_per_contract / days_to_expiry
            
            # Check for earnings risk (use cache to avoid repeated API calls)
            if ticker not in self.earnings_cache:
                self.earnings_cache[ticker] = self.provider.get_earnings_date(ticker)
            
            earnings_date = self.earnings_cache[ticker]
            earnings_risk = False
            earnings_days_diff = None
            has_earnings_before_expiry = False
            
            if earnings_date:
                earnings_days_diff = (earnings_date - expiry_date).days
                
                # Check if earnings falls between now and expiration
                if self.current_date < earnings_date <= expiry_date:
                    has_earnings_before_expiry = True
                    earnings_risk = True
                    
                    # GLOBAL EARNINGS EXCLUSION: Exclude ALL tickers if configured
                    if self.config.exclude_earnings_all:
                        return None  # Reject: earnings before expiration (all tickers)
                    
                    # HIGH-VOL STOCKS: Always exclude if earnings before expiry
                    if is_high_vol_stock and self.config.high_vol_exclude_earnings:
                        return None  # Reject: earnings before expiration (high-vol stocks)
                
                # Also flag as earnings risk if earnings is within 7 days after expiration
                # (market anticipation can affect option value)
                elif abs(earnings_days_diff) <= 7:
                    earnings_risk = True
            
            # Calculate break-even price
            if option_type == 'put':
                breakeven = strike - (premium_per_contract / 100)
            else:  # call
                breakeven = strike + (premium_per_contract / 100)
            
            # Get stock risk metrics (cached)
            if ticker not in self.risk_metrics_cache:
                self.risk_metrics_cache[ticker] = self.provider.get_stock_risk_metrics(ticker)
            
            risk_metrics = self.risk_metrics_cache[ticker]
            beta = risk_metrics['beta']
            market_cap = risk_metrics['market_cap']
            hist_vol = risk_metrics['hist_vol']
            atr_pct = risk_metrics['atr_pct']
            
            # Calculate IV rank (current IV vs 52-week range)
            # For now, use simple heuristic: IV > 40% = high, IV > 60% = very high
            if iv > 0:
                iv_pct = iv * 100
                if iv_pct > 60:
                    iv_rank_estimate = 80  # High IV environment
                elif iv_pct > 40:
                    iv_rank_estimate = 60  # Elevated IV
                elif iv_pct > 25:
                    iv_rank_estimate = 40  # Normal IV
                else:
                    iv_rank_estimate = 20  # Low IV
            else:
                iv_rank_estimate = 50  # Unknown
            
            # Classify stock as "boring" or "exciting"
            is_boring = (
                beta <= self.config.max_beta and
                hist_vol <= self.config.max_hist_vol and
                atr_pct <= self.config.max_atr_pct and
                market_cap >= self.config.min_market_cap
            )
            
            # Flag high-risk indicators
            risk_flags = []
            if beta > 1.5:
                risk_flags.append(f'HIGH_BETA_{beta:.1f}')
            if hist_vol > 0.45:
                risk_flags.append(f'HIGH_VOL_{hist_vol*100:.0f}%')
            if atr_pct > 4.0:
                risk_flags.append(f'HIGH_ATR_{atr_pct:.1f}%')
            if annualized_yield > self.config.max_annualized_for_boring:
                risk_flags.append(f'SUSPICIOUSLY_HIGH_YIELD_{annualized_yield*100:.0f}%')
            if market_cap < self.config.min_market_cap:
                risk_flags.append('SMALL_MID_CAP')
            if has_earnings_before_expiry:
                risk_flags.append('EARNINGS_BEFORE_EXPIRY')
            
            # FILTER OUT: Non-boring stocks with HIGH_BETA or HIGH_VOL (too risky)
            # These are volatile stocks with elevated risk - not suitable for boring wheel
            if not is_boring:
                has_high_beta = any('HIGH_BETA' in flag for flag in risk_flags)
                has_high_vol = any('HIGH_VOL' in flag for flag in risk_flags)
                if has_high_beta or has_high_vol:
                    return None  # Reject volatile non-boring stocks
            
            # OPPORTUNITY FLAG: Boring stock with suspiciously high yield = potential opportunity
            # Stable company with temporarily elevated premiums (market overreaction, fear, etc.)
            opportunity_flag = None
            if is_boring:
                has_high_yield = any('SUSPICIOUSLY_HIGH_YIELD' in flag for flag in risk_flags)
                if has_high_yield:
                    opportunity_flag = 'POSSIBLE_OPPORTUNITY'
            
            # In boring wheel mode, reject non-boring stocks
            if self.config.boring_wheel_mode and not is_boring:
                return None
            
            return {
                'ticker': ticker,
                'expiry': expiry_date.strftime('%Y-%m-%d'),
                'strike': strike,
                'option_type': option_type.upper(),
                'leg_type': leg_type,
                'underlying_price': underlying_price,
                'delta': delta,
                'premium': premium_per_contract,
                'bid': bid,
                'ask': ask,
                'mid': mid_price,
                'iv': iv * 100,  # Convert to percentage
                'roc': roc * 100,  # Percentage
                'annualized_yield': annualized_yield,  # Keep as decimal for scoring
                'pop': pop * 100,  # Percentage
                'spread_pct': spread_pct * 100,  # Percentage
                'cushion': cushion * 100,  # Percentage
                'rsi': technicals['rsi'],
                'adx': technicals['adx'],
                'collateral': collateral,
                'days_to_expiry': days_to_expiry,
                'ppd': ppd,
                'volume': volume,
                'open_interest': open_interest,
                'earnings_risk': earnings_risk,
                'has_earnings_before_expiry': has_earnings_before_expiry,
                'earnings_days_diff': earnings_days_diff if earnings_date else None,
                'dividend_risk': dividend_risk if option_type == 'put' else False,
                'breakeven': breakeven,
                'beta': beta,
                'market_cap': market_cap,
                'hist_vol': hist_vol * 100,  # Convert to percentage
                'atr_pct': atr_pct,
                'iv_rank_estimate': iv_rank_estimate,
                'is_boring': is_boring,
                'risk_flags': '|'.join(risk_flags) if risk_flags else None,
                'opportunity_flag': opportunity_flag,
            }
        
        except Exception as e:
            # Skip problematic options
            return None
    
    def scan_ticker(self, ticker: str, scan_csp: bool = True, scan_cc: bool = True, index: int = None, total: int = None) -> List[Dict]:
        """Scan all options for a single ticker"""
        
        if index is not None and total is not None:
            print(f"[{index}/{total}] Scanning {ticker}...", end=' ')
        else:
            print(f"Scanning {ticker}...", end=' ')
        
        try:
            # Get current price
            underlying_price = self.provider.get_current_price(ticker)
            if underlying_price <= 0:
                print("FAIL (no price)")
                return []
            
            # Get technical indicators
            technicals = self.calculate_technicals(ticker)
            
            # Get available expirations
            expirations = self.provider.get_expirations(ticker)
            
            if len(expirations) == 0:
                print("FAIL (no options)")
                return []
            
            candidates = []
            
            # Scan each expiration (limited to reduce API calls)
            for exp_str in expirations[:6]:  # Limit to first 6 expirations to avoid rate limiting
                try:
                    expiry_date = datetime.strptime(exp_str, '%Y-%m-%d')
                    
                    # Get option chain
                    calls, puts = self.provider.get_option_chain(ticker, exp_str)
                    
                    # Scan puts for CSP
                    if scan_csp and len(puts) > 0:
                        for idx, row in puts.iterrows():
                            result = self.analyze_option(
                                ticker, row, 'put', underlying_price, 
                                expiry_date, technicals
                            )
                            if result:
                                candidates.append(result)
                    
                    # Scan calls for CC
                    if scan_cc and len(calls) > 0:
                        for idx, row in calls.iterrows():
                            result = self.analyze_option(
                                ticker, row, 'call', underlying_price,
                                expiry_date, technicals
                            )
                            if result:
                                candidates.append(result)
                
                except Exception as e:
                    continue
            
            print(f"OK ({len(candidates)} candidates)")
            return candidates
        
        except Exception as e:
            print(f"FAIL ({str(e)})")
            return []
    
    def scan_universe(self, tickers: List[str], scan_csp: bool = True, 
                     scan_cc: bool = True) -> pd.DataFrame:
        """Scan entire universe of tickers"""
        
        print(f"\n{'='*80}")
        print(f"WHEEL STRATEGY SCANNER")
        print(f"Scanning {len(tickers)} tickers for wheel opportunities")
        print(f"Date: {self.current_date.strftime('%Y-%m-%d')}")
        print(f"{'='*80}\n")
        
        all_candidates = []
        
        for i, ticker in enumerate(tickers, 1):
            # Add small delay to avoid rate limiting (skip delay for first ticker)
            if i > 1:
                time.sleep(0.5)  # 500ms delay between tickers
            
            candidates = self.scan_ticker(ticker, scan_csp, scan_cc, index=i, total=len(tickers))
            all_candidates.extend(candidates)
        
        if len(all_candidates) == 0:
            print("\nNo candidates found matching criteria.")
            return pd.DataFrame()
        
        # Convert to DataFrame
        df = pd.DataFrame(all_candidates)
        
        # Calculate composite score with normalized metrics
        df = self._calculate_composite_score(df)
        
        # Sort by composite score descending
        df = df.sort_values('composite_score', ascending=False)
        
        return df
    
    def _calculate_composite_score(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate composite score with normalized metrics, weighted scoring, and BORING WHEEL adjustments.
        
        BASE Score = (2 × Annualized Yield Score) + (2 × Cushion Score) + 
                     Premium Yield Score + PoP Score + Liquidity Score
        
        BORING WHEEL Adjustments:
        - Cap annualized yield at 25% (higher yields = hidden risk, not feature)
        - ADD 2x stability bonus (low beta, low volatility, low ATR)
        - SUBTRACT 2x risk penalty (high beta, high volatility, suspiciously high yields)
        
        All metrics normalized 0-1, with emphasis on stability for true "boring" wheel strategy
        """
        if len(df) == 0:
            return df
        
        # Helper function to normalize using percentiles (more consistent across datasets)
        def normalize(series):
            # Use percentile ranking for consistent scores regardless of dataset size
            # This maps each value to its percentile (0-100), then scales to 0-1
            if len(series) == 0:
                return pd.Series([0.5] * len(series), index=series.index)
            
            # If all values are the same, return 0.5 for all (middle score)
            if series.std() == 0 or series.nunique() == 1:
                return pd.Series([0.5] * len(series), index=series.index)
            
            # For very small datasets (< 3 items), use min-max normalization instead of percentiles
            if len(series) < 3:
                min_val = series.min()
                max_val = series.max()
                if max_val == min_val:
                    return pd.Series([0.5] * len(series), index=series.index)
                return (series - min_val) / (max_val - min_val)
            
            # Calculate percentile for each value
            try:
                percentiles = series.apply(lambda x: percentileofscore(series, x, kind='rank'))
                # Scale from 0-100 to 0-1
                normalized = percentiles / 100.0
                # Replace any NaN values with 0.5 (middle score)
                return normalized.fillna(0.5)
            except Exception:
                # Fallback to min-max normalization if percentile fails
                min_val = series.min()
                max_val = series.max()
                if max_val == min_val:
                    return pd.Series([0.5] * len(series), index=series.index)
                return (series - min_val) / (max_val - min_val)
        
        # 1. Annualized Yield Score - CAP at 25% for boring wheel (anything higher = risk flag)
        df['annualized_capped'] = df['annualized_yield'].clip(upper=self.config.max_annualized_for_boring)
        df['annualized_score'] = normalize(df['annualized_capped'])
        
        # 2. Break-even Protection / Cushion Score (higher cushion = better) - DOUBLE WEIGHT
        df['cushion_score'] = normalize(df['cushion'])
        
        # 3. Premium Yield Score (higher ROC is better)
        df['premium_score'] = normalize(df['roc'])
        
        # 4. PoP Score (higher probability is better)
        df['pop_score'] = normalize(df['pop'])
        
        # 5. Liquidity Score (combine volume and OI, lower spread is better)
        # Normalize volume and OI (higher is better)
        df['volume_norm'] = normalize(df['volume'])
        df['oi_norm'] = normalize(df['open_interest'])
        # Normalize spread (lower is better, so invert)
        df['spread_norm'] = 1 - normalize(df['spread_pct'])
        # Liquidity score is average of volume, OI, and inverse spread
        df['liquidity_score'] = (df['volume_norm'] + df['oi_norm'] + df['spread_norm']) / 3
        
        # 6. STABILITY SCORE (lower beta + lower vol + lower ATR = more boring/stable)
        # Invert metrics so higher score = more stable
        df['beta_inv'] = df['beta'].apply(lambda x: max(0, 2.0 - max(0.5, min(x, 2.0))))  # Beta 0.5→1.5, Beta 2.0→0.0
        df['hist_vol_inv'] = df['hist_vol'].apply(lambda x: max(0, 60 - min(x, 60)))  # Low vol = high score
        df['atr_inv'] = df['atr_pct'].apply(lambda x: max(0, 5.0 - min(x, 5.0)))  # Low ATR = high score
        
        df['stability_raw'] = (df['beta_inv'] / 1.5 + df['hist_vol_inv'] / 60 + df['atr_inv'] / 5) / 3
        df['stability_score'] = normalize(df['stability_raw'])
        
        # 7. RISK PENALTY (penalize high yields >25%, high beta >1.5, high vol >40%, earnings before expiry)
        df['risk_penalty'] = 0.0
        df.loc[df['annualized_yield'] > 0.25, 'risk_penalty'] += 0.3  # -30% for suspiciously high yield
        df.loc[df['beta'] > 1.5, 'risk_penalty'] += 0.2  # -20% for high beta
        df.loc[df['hist_vol'] > 40, 'risk_penalty'] += 0.2  # -20% for high volatility  
        df.loc[df['atr_pct'] > 4.0, 'risk_penalty'] += 0.1  # -10% for high daily swings
        # Only apply earnings penalty if config flag is enabled
        if self.config.earnings_penalty_in_scoring:
            df.loc[df['has_earnings_before_expiry'] == True, 'risk_penalty'] += 0.4  # -40% for earnings risk (MAJOR penalty)
        
        # Calculate composite score: base metrics + stability bonus - risk penalty
        # Weights: 2x annualized (capped), 2x cushion, 1x premium, 1x PoP, 1x liquidity, 2x stability, -1x risk
        # Note: Reduced risk penalty weight from 2x to 1x to balance scoring
        df['composite_score'] = (
            2.0 * df['annualized_score'] +     # Double weight (but capped at 25%)
            2.0 * df['cushion_score'] +         # Double weight
            df['premium_score'] +
            df['pop_score'] +
            df['liquidity_score'] +
            2.0 * df['stability_score'] -       # BONUS for boring stocks (2x)
            1.0 * df['risk_penalty']            # PENALTY for exciting stocks (1x, reduced from 2x)
        )
        
        # Normalize composite score to 0-100 for easier interpretation
        # Max possible is 10.0 (2+2+1+1+1+2), min is 9.0 - 1.0 = 8.0 (if max risk penalty)
        df['composite_score'] = (df['composite_score'] / 10.0) * 100  # Scale to 0-100
        
        # Drop intermediate scoring columns to keep output clean
        df = df.drop(columns=['annualized_capped', 'annualized_score', 'cushion_score', 'premium_score', 
                              'pop_score', 'volume_norm', 'oi_norm', 'spread_norm', 
                              'liquidity_score', 'beta_inv', 'hist_vol_inv', 'atr_inv',
                              'stability_raw', 'stability_score', 'risk_penalty'])
        
        return df


def format_results(df: pd.DataFrame, max_rows: int = 50):
    """Format and display results"""
    
    if len(df) == 0:
        print("\nNo candidates found.")
        return
    
    print(f"\n{'='*150}")
    print(f"TOP {min(max_rows, len(df))} WHEEL STRATEGY CANDIDATES")
    print(f"{'='*150}\n")
    
    # Format columns for display
    display_df = df.head(max_rows).copy()
    
    # Round numeric columns
    display_df['delta'] = display_df['delta'].map('{:.3f}'.format)
    display_df['premium'] = display_df['premium'].map('${:.2f}'.format)
    display_df['iv'] = display_df['iv'].map('{:.1f}%'.format)
    display_df['roc'] = display_df['roc'].map('{:.2f}%'.format)
    display_df['annualized_yield'] = display_df['annualized_yield'].apply(lambda x: f'{x*100:.1f}%')  # Decimal to percent
    display_df['pop'] = display_df['pop'].map('{:.1f}%'.format)
    display_df['spread_pct'] = display_df['spread_pct'].map('{:.1f}%'.format)
    display_df['cushion'] = display_df['cushion'].map('{:.1f}%'.format)
    display_df['rsi'] = display_df['rsi'].map('{:.1f}'.format)
    display_df['adx'] = display_df['adx'].map('{:.1f}'.format)
    display_df['collateral'] = display_df['collateral'].map('${:,.0f}'.format)
    display_df['ppd'] = display_df['ppd'].map('${:.2f}'.format)
    display_df['underlying_price'] = display_df['underlying_price'].map('${:.2f}'.format)
    display_df['composite_score'] = display_df['composite_score'].map('{:.1f}'.format)
    display_df['beta'] = display_df['beta'].map('{:.2f}'.format)
    display_df['hist_vol'] = display_df['hist_vol'].map('{:.0f}%'.format)
    display_df['atr_pct'] = display_df['atr_pct'].map('{:.1f}%'.format)
    display_df['market_cap'] = display_df['market_cap'].apply(lambda x: f'${x/1e9:.1f}B' if x >= 1e9 else f'${x/1e6:.0f}M')
    
    # Select columns for display (add composite_score and risk metrics at beginning)
    cols = [
        'composite_score', 'ticker', 'leg_type', 'expiry', 'strike', 'delta', 'premium',
        'roc', 'annualized_yield', 'pop', 'cushion', 'spread_pct',
        'beta', 'hist_vol', 'atr_pct', 'market_cap', 'is_boring', 'opportunity_flag', 
        'has_earnings_before_expiry', 'risk_flags',
        'rsi', 'adx', 'collateral', 'days_to_expiry', 'ppd'
    ]
    
    print(display_df[cols].to_string(index=False))
    
    print(f"\n{'='*150}")
    print(f"SUMMARY STATISTICS")
    print(f"{'='*150}\n")
    
    print(f"Scoring: Composite Score (0-100) = 2×Annualized + 2×Cushion + ROC + PoP + Liquidity")
    print(f"Higher scores = better risk/reward balance\n")
    
    # Summary by leg type
    print("By Strategy Type:")
    print(df.groupby('leg_type').size())
    
    print("\nAverage Metrics:")
    numeric_cols = ['composite_score', 'roc', 'pop', 'cushion', 'days_to_expiry']
    print(df[numeric_cols].describe().loc[['mean', 'std', 'min', 'max']].round(2))
    
    # Show annualized yield separately (convert to percent for display)
    print("\nAnnualized Yield (%)")
    annualized_pct = df['annualized_yield'] * 100
    print(annualized_pct.describe().loc[['mean', 'std', 'min', 'max']].round(2))


def save_results(df: pd.DataFrame, filename: str = None):
    """Save results to Excel with highlighting"""
    
    if filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"wheel_candidates_{timestamp}.xlsx"
    
    # Ensure .xlsx extension
    if not filename.endswith('.xlsx'):
        filename = filename.replace('.csv', '.xlsx')
    
    # Save to OUTPUT folder
    import os
    output_dir = os.path.join(os.path.dirname(__file__), 'OUTPUT')
    os.makedirs(output_dir, exist_ok=True)
    filename = os.path.join(output_dir, os.path.basename(filename))
    
    # Create Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Candidates', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Candidates']
        
        # Define green fill for top candidates
        from openpyxl.styles import PatternFill
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        # Calculate score for highlighting (higher annualized_yield + higher pop + higher cushion)
        top_count = min(10, len(df))  # Top 10 or fewer
        
        # Highlight top rows (rows 2 to top_count+1, since row 1 is header)
        for row_idx in range(2, top_count + 2):
            for col_idx in range(1, len(df.columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = green_fill
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"\nResults saved to: {filename}")
    return filename


def create_detailed_report(df: pd.DataFrame, filename: str = None):
    """Create detailed markdown report for top 3 options"""
    
    if len(df) == 0:
        return
    
    if filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"wheel_top3_analysis_{timestamp}.md"
    
    # Save to OUTPUT folder
    import os
    output_dir = os.path.join(os.path.dirname(__file__), 'OUTPUT')
    os.makedirs(output_dir, exist_ok=True)
    filename = os.path.join(output_dir, os.path.basename(filename))
    
    top3 = df.head(3)
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write("# Top 3 Wheel Strategy Options - Detailed Analysis\n\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")
        f.write("**Note:** These options are ranked by a composite score that normalizes all metrics (0-1 scale) and gives double weight to Annualized Yield and Break-even Protection (Cushion).\n\n")
        f.write("---\n\n")
        
        # Show composite scores
        f.write("## Composite Scores (Ranked Best to Worst)\n\n")
        f.write("**Formula:** Score = (2 × Annualized Yield) + (2 × Cushion) + ROC + PoP + Liquidity\n\n")
        f.write("All metrics are normalized to 0-1 scale. Maximum possible score is 100.\n\n")
        
        for rank, (idx, row) in enumerate(top3.iterrows(), 1):
            f.write(f"**Rank {rank}: {row['ticker']} {row['expiry']} ${row['strike']:.2f}** - Score: {row['composite_score']:.1f}/100\n")
        f.write("\n---\n\n")
        
        # Detailed metric explanations
        f.write("## Metric Definitions\n\n")
        
        f.write("### Composite Score\n")
        f.write("**Definition:** A weighted scoring system that normalizes all metrics to a 0-1 scale and combines them. ")
        f.write("Annualized Yield and Cushion (break-even protection) get double weight because they most directly reflect risk/reward.\n\n")
        f.write("**Why it matters:** Prevents bias toward any single metric or ticker. A high annualized yield with no cushion scores lower than ")
        f.write("a moderate yield with good protection.\n\n")
        f.write("**Components:** 2×Annualized + 2×Cushion + ROC + PoP + Liquidity, normalized to 0-100.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** {row['composite_score']:.1f}/100\n")
        f.write("\n---\n\n")
        
        f.write("### ROC (Return on Capital)\n")
        f.write("**Definition:** The percentage return on the capital required to secure the position. ")
        f.write("For cash-secured puts, this is the premium received divided by the strike price × 100 (the cash collateral required).\n\n")
        f.write("**Why it matters:** Higher ROC means you're generating more income relative to the capital you're tying up. ")
        f.write("A 3% ROC on a 30-day position means you earn 3% on your capital in one month.\n\n")
        f.write("**Target:** Look for 2-5% for conservative positions, 5%+ for more aggressive trades.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** {row['roc']:.2f}%\n")
        f.write("\n---\n\n")
        
        f.write("### Annualized Yield (Double Weight in Score)\n")
        f.write("**Definition:** The ROC extrapolated to a full year based on trading days. Calculated as (ROC / DTE) × 252. ")
        f.write("This normalizes returns across different time periods using trading-day annualization (252 business days per year).\n\n")
        f.write("**Why it matters:** Allows you to compare a 21-day trade with a 45-day trade on equal footing. ")
        f.write("Shows the theoretical annual return if you could repeat this exact trade throughout the year. ")
        f.write("**Gets double weight in composite score.**\n\n")
        f.write("**Target:** 20-40% is conservative, 40-60% is moderate, 60%+ is aggressive.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** {row['annualized_yield']*100:.1f}%\n")
        f.write("\n---\n\n")
        
        f.write("### PoP (Probability of Profit)\n")
        f.write("**Definition:** Estimated probability that the option will expire OTM (out-of-the-money), meaning profitable for the seller. ")
        f.write("Calculated using Black-Scholes d2 probability approximation. This is a risk-neutral probability estimate based on implied volatility.\n\n")
        f.write("**Why it matters:** Higher PoP means the stock has more room to move against you before you lose money or get assigned. ")
        f.write("It's the statistical likelihood of keeping the full premium. Note: This does NOT account for earnings or gap risk.\n\n")
        f.write("**Target:** 80-85% is ideal for the wheel strategy - very high probability with reasonable premium.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** {row['pop']:.1f}%\n")
        f.write("\n---\n\n")
        
        f.write("### Cushion / Break-even Protection (Double Weight in Score)\n")
        f.write("**Definition:** The percentage distance between the current stock price and the strike price. ")
        f.write("Calculated as ((Current Price - Strike) / Current Price) × 100.\n\n")
        f.write("**Why it matters:** This is your safety buffer. A 5% cushion means the stock can drop 5% before you're at-the-money. ")
        f.write("Larger cushions mean lower risk but typically lower premiums. ")
        f.write("**Gets double weight in composite score as it directly measures risk protection.**\n\n")
        f.write("**Target:** 3-5% is conservative, 5-10% is very safe, <3% is aggressive.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** {row['cushion']:.1f}% ")
            f.write(f"(Stock: ${row['underlying_price']:.2f})\n")
        f.write("\n---\n\n")
        
        f.write("### Delta\n")
        f.write("**Definition:** Measures how much the option price changes for a $1 move in the stock. ")
        f.write("For puts, also approximates the probability of finishing in-the-money (ITM). Delta of -0.25 ≈ 25% chance of assignment.\n\n")
        f.write("**Why it matters:** Lower absolute delta = further out-of-the-money = safer but less premium. ")
        f.write("The wheel strategy typically uses deltas between -0.20 and -0.40 for good premium/risk balance.\n\n")
        f.write("**Target:** -0.20 to -0.30 is conservative, -0.30 to -0.40 is moderate risk.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** {row['delta']:.3f}\n")
        f.write("\n---\n\n")
        
        f.write("### Premium\n")
        f.write("**Definition:** The actual dollar amount you receive for selling one put contract (representing 100 shares).\n\n")
        f.write("**Why it matters:** This is the cash credit deposited into your account immediately. ")
        f.write("Higher premiums are attractive but usually come with higher risk.\n\n")
        f.write("**Target:** Varies by stock price, but aim for premiums that give 2-5% ROC.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** ${row['premium']:.2f}\n")
        f.write("\n---\n\n")
        
        f.write("### Days to Expiry (DTE)\n")
        f.write("**Definition:** Number of trading days (business days) until the option expires. Calculated using np.busday_count().\n\n")
        f.write("**Why it matters:** Shorter DTE means faster premium collection and more control (can adjust sooner), ")
        f.write("but requires more active management. Longer DTE means less frequent trading but capital is tied up longer.\n\n")
        f.write("**Target:** 21-45 trading days is the sweet spot for the wheel - balances premium decay with flexibility.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** {row['days_to_expiry']} days\n")
        f.write("\n---\n\n")
        
        f.write("### Spread %\n")
        f.write("**Definition:** The percentage difference between the bid and ask prices. ")
        f.write("Calculated as ((Ask - Bid) / Mid Price) × 100.\n\n")
        f.write("**Why it matters:** Wide spreads mean poor liquidity and you'll lose money on entry/exit. ")
        f.write("Narrow spreads (<5%) mean you can easily close the position early for profit.\n\n")
        f.write("**Target:** <5% is excellent, 5-10% is acceptable, >10% requires caution.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** {row['spread_pct']:.1f}%\n")
        f.write("\n---\n\n")
        
        f.write("### RSI (Relative Strength Index)\n")
        f.write("**Definition:** Technical indicator measuring recent price momentum on a scale of 0-100. ")
        f.write("Below 30 = oversold, above 70 = overbought.\n\n")
        f.write("**Why it matters:** For selling puts, lower RSI is better - the stock has pulled back and may bounce. ")
        f.write("Avoid selling puts when RSI > 80 (overbought) as the stock may be due for a pullback.\n\n")
        f.write("**Target:** 30-50 is ideal for selling puts, 50-70 is acceptable.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** {row['rsi']:.1f}\n")
        f.write("\n---\n\n")
        
        f.write("### Collateral Required\n")
        f.write("**Definition:** The cash you must have in your account to secure the put. ")
        f.write("Equals Strike Price × 100 shares.\n\n")
        f.write("**Why it matters:** This capital is locked up until expiration or until you close the position. ")
        f.write("Make sure you have this amount available and won't need it during the trade period.\n\n")
        f.write("**Target:** Should not exceed your risk limit per position (typically 10-25% of portfolio).\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** ${row['collateral']:,.0f}\n")
        f.write("\n---\n\n")
        
        f.write("### Premium Per Day (PPD)\n")
        f.write("**Definition:** Premium divided by days to expiry. Shows daily income rate.\n\n")
        f.write("**Why it matters:** Useful for comparing trades of different durations. ")
        f.write("Higher PPD means faster income generation.\n\n")
        f.write("**Target:** Varies by capital, but $10-50/day per contract is typical.\n\n")
        
        for idx, row in top3.iterrows():
            f.write(f"**{row['ticker']} {row['expiry']} ${row['strike']:.2f} Strike:** ${row['ppd']:.2f}/day\n")
        f.write("\n---\n\n")
        
        # Summary table
        f.write("## Top 3 Options Summary Table\n\n")
        f.write("| Rank | Ticker | Expiry | Strike | Premium | ROC | Annual | PoP | Cushion | DTE |\n")
        f.write("|------|--------|--------|--------|---------|-----|--------|-----|---------|-----|\n")
        
        for rank, (idx, row) in enumerate(top3.iterrows(), 1):
            f.write(f"| {rank} | {row['ticker']} | {row['expiry']} | ")
            f.write(f"${row['strike']:.2f} | ${row['premium']:.2f} | ")
            f.write(f"{row['roc']:.2f}% | {row['annualized_yield']*100:.1f}% | ")
            f.write(f"{row['pop']:.1f}% | {row['cushion']:.1f}% | {row['days_to_expiry']} |\n")
        
        f.write("\n---\n\n")
        f.write("## Action Items\n\n")
        f.write("1. **Verify liquidity** - Check volume and open interest before entering\n")
        f.write("2. **Check news** - Ensure no earnings or major events during the period\n")
        f.write("3. **Set alerts** - Monitor at 21 DTE and consider closing at 50% profit\n")
        f.write("4. **Position sizing** - Don't allocate more than 25% of portfolio to one ticker\n")
        f.write("5. **Exit plan** - Know your assignment strategy if put goes ITM\n")
    
    print(f"Detailed analysis saved to: {filename}")
    return filename


def should_run_now(config: WheelConfig) -> bool:
    """Return True if now is within the market open window on a weekday."""
    try:
        local_tz = ZoneInfo(config.user_timezone)
        now_local = datetime.now(local_tz)
        if now_local.weekday() >= 5:
            return False

        eastern_tz = ZoneInfo("America/New_York")
        now_eastern = now_local.astimezone(eastern_tz)
        market_open = now_eastern.replace(hour=9, minute=30, second=0, microsecond=0)
        window_end = market_open + timedelta(minutes=config.market_open_window_minutes)
        return market_open <= now_eastern <= window_end
    except Exception:
        return False


def get_top_boring_candidate(df: pd.DataFrame) -> Optional[pd.Series]:
    """Return the top boring candidate by composite score, or None if none exist."""
    if len(df) == 0 or 'is_boring' not in df.columns:
        return None
    boring_df = df[df['is_boring'] == True].copy()
    if len(boring_df) == 0:
        return None
    boring_df = boring_df.sort_values('composite_score', ascending=False)
    return boring_df.iloc[0]


def build_go_no_go_statement(df: pd.DataFrame, config: WheelConfig) -> str:
    """Build GO/NO GO statement for the top boring candidate."""
    top = get_top_boring_candidate(df)
    if top is None:
        return "NO GO - No boring wheel candidates passed filters today. Better to wait for a more optimal day to trade."

    score = float(top['composite_score'])
    pop = float(top['pop'])
    cushion = float(top['cushion']) / 100.0
    annualized = float(top['annualized_yield'])

    meets = (
        score >= config.go_min_composite_score and
        pop >= config.go_min_pop and
        cushion >= config.go_min_cushion and
        annualized <= config.max_annualized_for_boring
    )

    ticker = top['ticker']
    expiry = top['expiry']
    strike = float(top['strike'])
    pop_pct = pop * 100.0
    cushion_pct = float(top['cushion'])

    if meets:
        return (
            f"GO - Top boring wheel candidate {ticker} {expiry} ${strike:.2f} "
            f"scores {score:.1f} with {pop_pct:.0f}% PoP and {cushion_pct:.1f}% cushion. "
            "Looks worthy of a trade."
        )

    return (
        f"NO GO - Top boring wheel candidate {ticker} {expiry} ${strike:.2f} "
        f"scores {score:.1f} with {pop_pct:.0f}% PoP and {cushion_pct:.1f}% cushion, "
        "but does not meet thresholds. Better to wait for a more optimal day to trade."
    )


def send_email_with_attachments(subject: str, body: str, attachments: List[str], config: WheelConfig) -> None:
    """Send an email with attachments using SMTP settings from environment variables."""
    smtp_host = os.getenv("EMAIL_HOST", "")
    smtp_port = int(os.getenv("EMAIL_PORT", "465"))
    smtp_user = os.getenv("EMAIL_USER", "")
    smtp_pass = os.getenv("EMAIL_PASS", "")
    from_addr = os.getenv("EMAIL_FROM", smtp_user)
    to_addr = os.getenv("EMAIL_TO", "TRADCLIMBER@GMAIL.COM")

    if not smtp_host or not smtp_user or not smtp_pass or not from_addr:
        print("Email not sent: missing EMAIL_HOST/EMAIL_USER/EMAIL_PASS/EMAIL_FROM settings.")
        return

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = to_addr
    msg.set_content(body)

    for path in attachments:
        if not path or not os.path.exists(path):
            continue
        filename = os.path.basename(path)
        if filename.lower().endswith(".xlsx"):
            maintype = "application"
            subtype = "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif filename.lower().endswith(".md"):
            maintype = "text"
            subtype = "markdown"
        else:
            maintype = "application"
            subtype = "octet-stream"

        with open(path, "rb") as f:
            msg.add_attachment(f.read(), maintype=maintype, subtype=subtype, filename=filename)

    if smtp_port == 465:
        with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
    else:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)


def get_default_universe() -> List[str]:
    """Return expanded ticker universe for scanning - 300+ stocks across all sectors (deduplicated)"""
    
    tickers = [
        # Large cap tech (Mega cap + Growth)
        'AAPL', 'MSFT', 'GOOGL', 'GOOG', 'META', 'NVDA', 'AMD', 'INTC', 'CSCO', 'AVGO', 
        'ORCL', 'CRM', 'ADBE', 'IBM', 'TXN', 'QCOM', 'AMAT', 'LRCX', 'KLAC', 'MU',
        'ADI', 'MRVL', 'NXPI', 'SNPS', 'CDNS', 'PANW', 'CRWD', 'ZS', 'DDOG', 'NET',
        'PLTR', 'ANET', 'FTNT', 'WDAY', 'NOW', 'TEAM', 'ZM', 'DOCU',
        
        # Finance & Banking
        'JPM', 'BAC', 'WFC', 'C', 'GS', 'MS', 'BLK', 'SCHW', 'AXP', 'USB',
        'PNC', 'TFC', 'COF', 'BK', 'STT', 'FITB', 'KEY', 'RF', 'CFG', 'HBAN',
        'CME', 'ICE', 'SPGI', 'MCO', 'COIN', 'HOOD', 'SOFI', 'AFRM', 'PYPL', 'SQ',
        
        # Insurance & Real Estate
        'BRK.B', 'PGR', 'TRV', 'ALL', 'AIG', 'MET', 'PRU', 'AFL', 'HIG', 'CB',
        'PLD', 'AMT', 'EQIX', 'PSA', 'SPG', 'O', 'WELL', 'DLR', 'AVB', 'EQR',
        
        # Consumer Discretionary
        'AMZN', 'TSLA', 'HD', 'LOW', 'MCD', 'SBUX', 'NKE', 'TJX', 'BKNG', 'MAR',
        'ABNB', 'LULU', 'DECK', 'RCL', 'CCL', 'NCLH', 'LVS', 'WYNN', 'MGM', 'CMG',
        'YUM', 'QSR', 'DPZ', 'DKNG', 'PENN', 'F', 'GM', 'RIVN', 'LCID', 'UBER',
        'LYFT', 'DASH', 'EXPE', 'EBAY', 'ETSY', 'W', 'CHWY', 'CVNA', 'KMX',
        
        # Consumer Staples (Boring wheel favorites!)
        'WMT', 'COST', 'PG', 'KO', 'PEP', 'PM', 'MO', 'CL', 'KMB', 'GIS',
        'K', 'CPB', 'KHC', 'HSY', 'MDLZ', 'STZ', 'TAP', 'BUD', 'KDP', 'MNST',
        'KR', 'SYY', 'TSN', 'HRL', 'CAG', 'EL', 'CLX', 'CHD', 'CVS',
        
        # Healthcare & Pharma
        'UNH', 'JNJ', 'LLY', 'ABBV', 'MRK', 'PFE', 'TMO', 'ABT', 'DHR', 'BMY',
        'AMGN', 'GILD', 'VRTX', 'REGN', 'BIIB', 'MRNA', 'BNTX', 'ZTS', 'ISRG', 'SYK',
        'BSX', 'MDT', 'EW', 'IDXX', 'DXCM', 'ALGN', 'HCA', 'UHS', 'DGX', 'LH',
        'CI', 'HUM', 'CNC', 'MOH', 'ELV', 'CAH', 'MCK', 'VEEV',
        
        # Energy & Utilities (Boring wheel favorites!)
        'XOM', 'CVX', 'COP', 'SLB', 'EOG', 'PSX', 'MPC', 'VLO', 'OXY', 'HAL',
        'BKR', 'NOV', 'DVN', 'FANG', 'APA', 'KMI', 'WMB',
        'NEE', 'DUK', 'SO', 'D', 'AEP', 'EXC', 'SRE', 'PEG', 'XEL', 'ED',
        
        # Industrials & Aerospace
        'BA', 'CAT', 'GE', 'HON', 'UPS', 'FDX', 'UNP', 'CSX', 'NSC', 'LUV',
        'DAL', 'AAL', 'UAL', 'RTX', 'LMT', 'NOC', 'GD', 'LHX', 'TDG', 'HWM',
        'DE', 'EMR', 'ITW', 'ETN', 'PH', 'ROK', 'AME', 'DOV', 'IR', 'CARR',
        
        # Materials & Chemicals
        'LIN', 'APD', 'ECL', 'SHW', 'DD', 'DOW', 'PPG', 'NEM', 'FCX', 'GOLD',
        'NUE', 'STLD', 'RS', 'VMC', 'MLM', 'ALB', 'FMC', 'CE', 'CF', 'MOS',
        
        # Telecom & Media
        'T', 'VZ', 'TMUS', 'CMCSA', 'DIS', 'NFLX', 'WBD', 'FOXA', 'SPOT',
        'RBLX', 'EA', 'TTWO', 'U', 'MTCH', 'BMBL', 'SNAP', 'PINS', 'TWLO',
        
        # Semiconductors & Hardware
        'TSM', 'ASML', 'MPWR', 'ON', 'SWKS', 'MCHP', 'TER',
        
        # Retail & E-commerce
        'TGT', 'DG', 'DLTR', 'ROST', 'BBWI', 'GPS',
        'AEO', 'ANF', 'FL', 'DKS', 'BBY', 'FIVE', 'OLLI', 'BJ', 'BIG', 'TSCO',
        
        # International & Emerging
        'BABA', 'JD', 'BIDU', 'NIO', 'XPEV', 'LI', 'TME', 'BILI',
        
        # ETFs (Broad market + Sector)
        'SPY', 'QQQ', 'IWM', 'TLT', 
        
        # REITs & Infrastructure
        'VTR', 'ARE', 'INVH', 'MAA', 'ESS', 'UDR', 'CPT', 'AIV', 'PEAK', 'IRM'
    ]
    
    # Deduplicate and return
    return list(dict.fromkeys(tickers))  # Preserves order while removing duplicates


def main():
    """Main execution"""
    parser = argparse.ArgumentParser(description="Wheel Strategy Scanner")
    parser.add_argument(
        "--run-now",
        action="store_true",
        help="Bypass the market-open schedule guard"
    )
    args = parser.parse_args()

    # Initialize data provider
    provider = YFinanceProvider()
    
    # Initialize config with default settings
    config = WheelConfig()

    if not args.run_now and not should_run_now(config):
        print("Outside market open window. Use --run-now to bypass.")
        return
    
    # Customize config if needed
    # config.min_dte = 25
    # config.max_dte = 45
    # config.min_annualized = 0.20  # 20% minimum
    
    # Initialize scanner
    scanner = WheelScanner(provider, config)
    
    # Get ticker universe
    # Option 1: Use default universe
    tickers = get_default_universe()
    
    # Option 2: Custom list
    # tickers = ['KO', 'PEP', 'WMT', 'PG', 'T']
    
    # Option 3: Focus on your current position
    # tickers = ['KO']
    
    # Scan universe
    # Set scan_csp=True to find cash-secured put opportunities
    # Set scan_cc=True to find covered call opportunities
    results = scanner.scan_universe(
        tickers,  # Full universe
        scan_csp=True,  # Scan for CSP candidates
        scan_cc=False   # Don't scan CC (unless you own shares)
    )
    
    # Display results
    format_results(results, max_rows=30)
    
    # Save to Excel with highlighting and create detailed report
    if len(results) > 0:
        excel_path = save_results(results)  # Saves as Excel with green highlighting
        md_path = create_detailed_report(results)  # Creates markdown report for top 3

        go_no_go = build_go_no_go_statement(results, config)
        md_body = ""
        if md_path and os.path.exists(md_path):
            with open(md_path, "r", encoding="utf-8") as f:
                md_body = f.read()
        else:
            md_body = "No markdown report was generated."

        email_body = f"{go_no_go}\n\n{md_body}"
        subject = f"Wheel Scanner Results {datetime.now().strftime('%Y-%m-%d')}"
        send_email_with_attachments(subject, email_body, [excel_path, md_path], config)
    else:
        go_no_go = build_go_no_go_statement(results, config)
        email_body = f"{go_no_go}\n\nNo candidates found."
        subject = f"Wheel Scanner Results {datetime.now().strftime('%Y-%m-%d')}"
        send_email_with_attachments(subject, email_body, [], config)


if __name__ == "__main__":
    main()
